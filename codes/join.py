#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

################################################
#
# 2022-10-21 处理北京
#
# 2022-10-28 处理中高两个文件
#
################################################


import openpyxl
import os
import re
import datetime

def join():
    
    # 23 + 5 + 4 + 2
    province = [ '河北省', '山西省', '辽宁省', '吉林省', 
                 '黑龙江省', '江苏省', '浙江省', '安徽省',
                 '福建省', '江西省', '山东省', '河南省', 
                 '湖北省', '湖南省', '广东省', '海南省', 
                 '四川省', '贵州省', '云南省', '陕西省', 
                 '甘肃省', '青海省', '台湾省',
                 '内蒙古自治区', '广西壮族自治区', 
                 '西藏自治区', '宁夏回族自治区', 
                 '新疆维吾尔自治区',
                 '香港特别行政区', '澳门特别行政区' ]
    
    municipality = {'天津市', '上海市', '重庆市'}
    # 当前是重庆郊县发生疫情，重庆市暂时不用判断
    
    # normal
    capital = { '北京市' }
    
    txt_arr = [ ['高风险', 'high.txt'],
                ['中风险', 'mid.txt'] ]
    
    for tid in range(len(txt_arr)):
    
        xlsx_path = r'../data/template_risk.xlsx'
        fn_path = '../data/src/fxq/%s' %(txt_arr[tid][1])
        
        fn_list = []
        stack = [] #风险区栈 拼接省-市-县
        pattern = r'(.+)\(\d+\)个'
        
    
        # open file
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb['Sheet1']
        rid = 2 # sheet record id
        
        with open(fn_path, 'r') as f:
            fn_list = f.readlines()
        
            #print(fn_list)
            
        for value in fn_list:
            value = value.strip('\n')
            ret = re.search(pattern, value)
            if ret: # 1st if : 处理带括号的数据
                # 判断省和市
                key = ret.group(1)
                if key in province: # 2nd if : 判断省
                    if len(stack)==0: # 3rd if : 判断是否到了下一省份
                        #print(key)
                        stack.append(key)
                    else:
                        #到了下一省份
                        if len(stack) != 2:
                            print('stack Error!%s' %(stack) )
                            break
                        stack.pop() #弹出市
                        stack.pop() #弹出省
                        stack.append(key) #new province in stack
                elif key in municipality: # 2nd if : 判断直辖市
                    if len(stack)==0:
                        #print(key)
                        stack.append(key)
                        stack.append(key)
                    else:
                        #到了下一省份
                        if len(stack) != 2:
                            print('stack Error!%s' %(stack) )
                            break
                        stack.pop() #弹出市
                        stack.pop() #弹出省
                        stack.append(key) #new province in stack
                        stack.append(key) #new province in stack
                    
                elif key in capital: # 2nd if : 判断直辖市
                    if len(stack)==0:
                        #print(key)
                        stack.append(key)
                        stack.append(key)
                    else:
                        #到了下一省份
                        if len(stack) != 2:
                            print('stack Error!%s' %(stack) )
                            break
                        stack.pop() #弹出市
                        stack.pop() #弹出省
                        stack.append(key) 
                        stack.append(key) 
                else: # 这是市级行政区
                    if len(stack) == 2 and stack[-2] not in capital:
                        stack.pop()
                    #print('\t%s' %(key) )
                    
                    if stack[-1] not in capital:
                        stack.append(key)
                    else:
                        county = key
                        #print('key is %s' %(key))
                continue
            #print(value)
            
        
            rid = rid+1
            sheet.cell(rid, 1).value = txt_arr[tid][0] # col A
            sheet.cell(rid, 2).value = stack[-2] # col B
            sheet.cell(rid, 3).value = stack[-1] # col C
            if stack[-2] in capital:
                sheet.cell(rid, 4).value = county # col D
                #print('%s\t%s\t%s' %(stack[-2], stack[-1], county) )
            else:
                sheet.cell(rid, 4).value = value # col D
                #print('%s\t%s\t%s' %(stack[-2], stack[-1], value) )


        # final delete duplicate lines of excel
        # in fact, only Beijing. it is the problem of history.
        wb_del = openpyxl.load_workbook(xlsx_path)
        sheet_del = wb_del['Sheet1']
        rid = 2 # sheet record id

        for i in range(3, sheet.max_row+1):

            if sheet.cell(i, 2).value == sheet.cell(i-1,2).value and sheet.cell(i, 3).value == sheet.cell(i-1,3).value and sheet.cell(i, 4).value == sheet.cell(i-1,4).value:
                #                print('\t%s\t%s\t%s\t%s' %(sheet.cell(i,1).value,
                #                                         sheet.cell(i,2).value,
                #                                         sheet.cell(i,3).value,
                #                                         sheet.cell(i,4).value ))
                continue

            rid = rid + 1
            sheet_del.cell(rid, 1).value = sheet.cell(i, 1).value
            sheet_del.cell(rid, 2).value = sheet.cell(i, 2).value
            sheet_del.cell(rid, 3).value = sheet.cell(i, 3).value
            sheet_del.cell(rid, 4).value = sheet.cell(i, 4).value
            
            
    
        today = datetime.datetime.now()
        dst_fn = '%s_%04d-%02d-%02d.xlsx' %( txt_arr[tid][0],
                                             today.year,
                                             today.month,
                                             today.day)
        wb_del.save('../data/dst/%s' %(dst_fn) )
        print('\n\t Delete duplicate %d lines.' %(sheet.max_row-rid) )
        print('%s has generated. \tTotal %d records.' %(dst_fn, rid-2) )
