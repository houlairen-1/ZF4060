#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

import openpyxl
import os
import re
import logging


xlsx_path = r'../data/src/fxq_src.xlsx'
fn_path = '../data/dst/fxq.txt'
log_path = '../fxq.log'

log = open(log_path, encoding='utf-8', mode='a')
logging.basicConfig(level=logging.ERROR,
                    stream = log,
                    format="%(asctime)s - %(name)s - %(levelname)-9s - %(filename)-8s : %(lineno)s line - %(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S"
                    )


# open file
municipality = {'北京市', '天津市', '上海市', '重庆市'}

wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Sheet1']

with open(fn_path, 'w') as f:
    rows = sheet.max_row
    
    for rid in range(2,rows+1):
        province = sheet.cell(rid,1).value
        if province:
            cur_province = province
        city = sheet.cell(rid,2).value
        county = sheet.cell(rid,3).value
        if cur_province in municipality:
            city = cur_province
            county = sheet.cell(rid,2).value
        # split column
        if not county:
            print('[ERROR]county is none:{}\t{}\t{}\n'.format(cur_province, city, county))
            logging.error('county is none:{}\t{}\t{}\n'.format(cur_province, city, county))
            continue

        # why do many splits exist ?
        # Jining lacked rules. 中文字符和英文字符都来一遍吧！
        county = county.replace('；','\n\t\t\t')
        county = county.replace(';','\n\t\t\t')
        county = county.replace('、','\n\t\t\t')
        county = county.replace('，','\n\t\t\t')
        county = county.replace(',','\n\t\t\t')
        #print('中风险\t%s\t%s\t%s\n' 
        #      %( cur_province, city, county ) )

        f.write('中风险\t%s\t%s\t%s\n' 
                %( cur_province, city, county ) )

