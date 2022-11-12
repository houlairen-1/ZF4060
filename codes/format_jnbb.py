#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

import openpyxl

src_xlsx_path = r'../data/src/jnbb_src.xlsx'
dst_xlsx_path = r'../data/bak.xlsx'

wb_src = openpyxl.load_workbook(src_xlsx_path)
sheet_src = wb_src.active
wb_dst = openpyxl.load_workbook(dst_xlsx_path)
sheet_dst = wb_dst.active

# get shape of src
# start from 4th row
rows = sheet_src.max_row
#print(rows)

key_dict = { '赵固堆派出所':'赵堌堆乡', 
             '馆驿派出所':'馆驿镇' }

for rid in range(4,rows+1):
    # 2022/10/18 modify in jnbb_src.
    # so suit to it.
    for cid in range(1, 21): #col: A-T total 20 columns
        sheet_dst.cell(rid,cid).value = sheet_src.cell(rid,cid).value
        #print('%d\t%s' %(rid, sheet_src.cell(rid,cid).value))
    sheet_dst.cell(rid,29).value = sheet_src.cell(rid,21).value
    
    for cid in range(23, 27): #col: A-L
        sheet_dst.cell(rid,cid-2).value = sheet_src.cell(rid,cid).value

wb_dst.save('../data/dst/jnbb.xlsx')
