#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

import openpyxl
import os

src_path = '../data/src/sf/'
xlfs = [x for x in os.listdir(src_path) if os.path.splitext(x)[1] == '.xlsx'] # 罗列目录内所有xlsx文件
print('需要统计',len(xlfs) , '个表格')

xl0 = xlfs[0]
data0 = []#复制表头数据
wb0 = openpyxl.load_workbook(filename = src_path+xl0)
ws0 = wb0.active
#print(ws0.max_column)
for i in range(1,ws0.max_column+1):
    data0.append(ws0.cell(row = 1,column = i).value)
#print('表头',data0)

data1 = []#复制数据
num = len(xlfs)
for n in range(num):
    xf = xlfs[n]
    wb1 = openpyxl.load_workbook(filename = src_path+xf)
    ws1 = wb1.active
    for i in range(2,ws1.max_row + 1):
        list = []
        for j in range(1,ws1.max_column + 1):
            list.append(ws1.cell(row=i,column=j).value)
        data1.append(list)
#print('数据',data1)

# # 汇总表头和数据,新建保存总表
data=[]
data.append(data0)#添加表头
for l in range(len(data1)):#添加数据
    data.append(data1[l])
wb = openpyxl.Workbook()#新建表
ws = wb.active
ws.title = '汇总'
for n_row in range(1,len(data)+1):#写入数据
    for n_col in range(1,len(data[n_row-1])+1):
        ws.cell(row=n_row,column=n_col,value=str(data[n_row-1][n_col-1]))
wb.save(filename='../data/tmp_sf.xlsx')#保存xlsx
print ('汇总完成')


src_xlsx_path = r'../data/tmp_sf.xlsx'
dst_xlsx_path = r'../data/bak.xlsx' # this is template of upload

key_dict = { '赵固堆派出所':'赵堌堆乡', 
             '馆驿派出所':'馆驿镇', 
             '水泊派出所':'水泊街道', 
             '大路口派出所':'大路口乡', 
             '韩岗派出所':'韩岗镇', 
             '安民派出所':'梁山街道', 
             '杨营派出所':'杨营镇', 
             '韩垓派出所':'韩垓镇', 
             '黑虎庙派出所':'黑虎庙镇', 
             '寿张集派出所':'寿张集镇', 
             '拳铺派出所':'拳铺镇', 
             '杏花村派出所':'水泊街道', 
             '徐集派出所':'拳铺镇', 
             '小安山派出所':'小安山镇', 
             '小路口派出所':'小路口镇',
             '马营派出所':'马营镇',
             '凤凰山派出所':'梁山街道' }

wb_src = openpyxl.load_workbook(src_xlsx_path)
sheet_src = wb_src.active
wb_dst = openpyxl.load_workbook(dst_xlsx_path)
sheet_dst = wb_dst.active

# get shape of src
# start from 2th row
rows = sheet_src.max_row
#print(rows)

# dst xlsx starts from 4
# template id is tid
tid = 3
for rid in range(2,rows+1):
    # 判断所属镇街 G7 ==> A1
    find_flag = False
    value = sheet_src.cell(rid,17).value

    for k, v in key_dict.items():
        if k in value:
            tid = tid+1
            find_flag = True
            sheet_dst.cell(tid,1).value = v
            break
    
    if not find_flag:
        continue

    sheet_dst.cell(tid,4).value = sheet_src.cell(rid,1).value
    sheet_dst.cell(tid,6).value = sheet_src.cell(rid,2).value
    sheet_dst.cell(tid,7).value = sheet_src.cell(rid,3).value
    sheet_dst.cell(tid,8).value = sheet_src.cell(rid,8).value
    sheet_dst.cell(tid,9).value = sheet_src.cell(rid,9).value
    sheet_dst.cell(tid,18).value = sheet_src.cell(rid,10).value
    sheet_dst.cell(tid,17).value = sheet_src.cell(rid,11).value
    sheet_dst.cell(tid,23).value = sheet_src.cell(rid,12).value # 最新核算检测时间
    sheet_dst.cell(tid,27).value = sheet_src.cell(rid,13).value

    
    

wb_dst.save('../data/dst/sf.xlsx')
