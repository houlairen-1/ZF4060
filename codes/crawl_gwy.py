#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.chrome.service import Service
import re
import openpyxl
import datetime
from time import sleep


MAX_PAGES = 100
NUM_BUTTONS = 9

#browser = webdriver.Chrome(executable_path="../drivers/chromedriver") #mac
# browser = webdriver.Chrome(executable_path="../drivers/chromedriver.exe") #windows
# Warning:  DeprecationWarning: executable_path has been deprecated, please pass in a Service object 版本过时
options = webdriver.ChromeOptions()
# solve SSL 证书错误问题
options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')

# ignore useless log
options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])

#ser = Service("../drivers/chromedriver.exe") #windows
ser = Service("../drivers/chromedriver") # macos
browser = webdriver.Chrome(service=ser, options=options)
browser.get('https://bmfw.www.gov.cn/yqfxdjcx/risk.html')
browser.implicitly_wait(5)
#Wait(browser, 30).until(
#    Expect.presence_of_element_located((By.CLASS_NAME, "risk-info-table"))
#)

risk_levels_list = [ '高风险', '低风险' ]

for level in risk_levels_list:
    f = open('../data/src/fxq/crawl_%s.txt' %(level), 'w')

    # 等待加载
    xlsx_path = r'../data/template_risk.xlsx'
    f.write('Flag:\t%s\n' %(level) )
    # open file
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb['Sheet1']
    rid = 2 # sheet record id
    
    # eg.中风险区 （2155）
    #level_pattern = ' %s区 （\d+）' %(level)
    level_pattern = ' %s区 .*' %(level)
    risk_levels = browser.find_elements(By.CLASS_NAME, 'tabs-header-tab')
    find_level = False
    for lid in range( len(risk_levels) ):
        risk_level = risk_levels[lid].get_attribute('textContent')
        if re.search(level_pattern, risk_level):
            find_level = True
            risk_levels[lid].click()
            #Wait(browser, 60).until(
            #    Expect.presence_of_element_located((By.CLASS_NAME, "loading"))
            #)
            print('%s has found.\t Click it.' %(level))
            break

    if not find_level:
        print('Level %s doesnot have found.' %(level) )
        continue


    # visit all pages
    for pid in range(1, MAX_PAGES):
        sleep(2)
        btn = browser.find_elements(By.CSS_SELECTOR, 'div.pages-box > button')
        # find pid in 9 buttons
        # bid: 0 is 首页, 1 is 上一页
        find_flag = False
        for bid in range( 2, NUM_BUTTONS-2 ):
            page_info = btn[bid].get_attribute('textContent')
            if int(page_info) == pid:
                find_flag = True
                print( 'Page %02d:\t crawl info.' %(pid) )
                break
    
        if not find_flag:
            print('Total pages is %02d.' %(pid-1) )
            break
    
        # next page
        btn[bid].click()
        
        risk_table = browser.find_elements(By.CLASS_NAME, "risk-info-table")
        for risk_title in risk_table:
            # <div>山西省 晋城市 城区</div>
            risk_row = risk_title.find_element(By.CLASS_NAME, "risk-info-table-title")
            #record_txt = risk_row.get_attribute("textContent")
            #print(risk_row.get_attribute("textContent"))
            #record = risk_row.get_attribute("textContent")
            record_html = risk_row.get_attribute("innerHTML")
            record_pattern = '<div>(.*)</div><div.*>(.*)</div>'
            try:
                record_txt = re.match(record_pattern, record_html).group(1)
            except:
                print('[ERROR] Pattern does not match.\t%s' %(risk_row.get_attribute("textContent")))
                continue
            #print(record_txt)
            f.write('%s\n' %(record_txt))
            record_arr = record_txt.split(' ')
            rid = rid+1
            sheet.cell(rid, 1).value = level
            for tid in range(len(record_arr)):
                sheet.cell(rid, tid+2).value = record_arr[tid]
            # single page
            
    
    
    if pid == MAX_PAGES:
        print('MAX_PAGES is smaller.')

    f.close()
    today = datetime.datetime.now()
    dst_fn = '%s_%04d-%02d-%02d.xlsx' %( level,
                                         today.year,
                                         today.month,
                                         today.day)

    wb.save('../data/dst/%s' %(dst_fn))
    print('%s has generated.' %(dst_fn))
