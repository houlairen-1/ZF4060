#!/usr/local/bin/python3
#_*_ coding:utf-8 _*_

import openpyxl
import os

src_path = '../data/src/'

xlfs = [x for x in os.listdir(src_path) if os.path.splitext(x)[1] == '.xlsx'] # 罗列目录内所有xlsx文件
print('需要统计',len(xlfs) , '个表格')

xl0 = xlfs[0]
data0 = []#复制表头数据
wb0 = openpyxl.load_workbook(filename = src_path+xl0)
ws0 = wb0.active
for i in range(1,ws0.max_column+1):
    data0.append(ws0.cell(row = 1,column = i).value)
#print('表头',data0)

data1 = []#复制数据
num = len(xlfs)
for n in range(num):
    xf = xlfs[n]
    wb1 = openpyxl.load_workbook(filename = src_path+xf)
    ws1 = wb1.active
    for i in range(2,ws1.max_row + 1):
        list = []
        for j in range(1,ws1.max_column + 1):
            list.append(ws1.cell(row=i,column=j).value)
        data1.append(list)
#print('数据',data1)

# # 汇总表头和数据,新建保存总表
data=[]
data.append(data0)#添加表头
for l in range(len(data1)):#添加数据
    data.append(data1[l])
wb = openpyxl.Workbook()#新建表
ws = wb.active
ws.title = '汇总'
for n_row in range(1,len(data)+1):#写入数据
    for n_col in range(1,len(data[n_row-1])+1):
        ws.cell(row=n_row,column=n_col,value=str(data[n_row-1][n_col-1]))
wb.save(filename='../data/a总表.xlsx')#保存xlsx
print ('汇总完成')


src_xlsx_path = r'../data/a总表.xlsx'
dst_xlsx_path = r'../data/bak.xlsx'

wb_src = openpyxl.load_workbook(src_xlsx_path)
sheet_src = wb_src.active
wb_dst = openpyxl.load_workbook(dst_xlsx_path)
sheet_dst = wb_dst.active

# get shape of src
# start from 2th row
rows = sheet_src.max_row
#print(rows)

# dst xlsx starts from 4
for rid in range(2,rows+1):
    sheet_dst.cell(rid+2,17).value = sheet_src.cell(rid,1).value
    sheet_dst.cell(rid+2,7).value = sheet_src.cell(rid,2).value
    sheet_dst.cell(rid+2,1).value = sheet_src.cell(rid,5).value
    sheet_dst.cell(rid+2,23).value = sheet_src.cell(rid,13).value
    sheet_dst.cell(rid+2,4).value = sheet_src.cell(rid,18).value
    sheet_dst.cell(rid+2,6).value = sheet_src.cell(rid,19).value
    sheet_dst.cell(rid+2,24).value = sheet_src.cell(rid,21).value
    sheet_dst.cell(rid+2,8).value = sheet_src.cell(rid,28).value
    sheet_dst.cell(rid+2,9).value = sheet_src.cell(rid,29).value
    sheet_dst.cell(rid+2,11).value = sheet_src.cell(rid,30).value
    sheet_dst.cell(rid+2,12).value = sheet_src.cell(rid,31).value
    sheet_dst.cell(rid+2,18).value = sheet_src.cell(rid,32).value

wb_dst.save('../data/ret-电话排查.xlsx')
